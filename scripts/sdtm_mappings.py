from __future__ import annotations

from collections import defaultdict
from dataclasses import dataclass
import html
from pathlib import Path
import re

try:
    from linkml_runtime.utils.schemaview import SchemaView
except ImportError:  # pragma: no cover - handled by skipping class-page injection
    SchemaView = None

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover - handled with a placeholder page
    load_workbook = None


DOMAIN_PAGES_DIRNAME = "sdtm-mappings"
CLASS_DOCS_DIRNAME = "classes"
DETAILS_PAGE_SUFFIX = "-details"
SECTION_METADATA: dict[str, tuple[str, str]] = {
    "TA": (
        "Trial Arms",
        "Arm-level trial design content, including planned arm descriptions, arm codes, and arm-specific transitions between epochs and elements.",
    ),
    "TE": (
        "Trial Elements",
        "Element-level trial design content, including element names, descriptions, durations, and transition rules.",
    ),
    "TV": (
        "Trial Visits",
        "Visit-level mappings for encounter names, visit order, planned day logic, and visit rules.",
    ),
    "TI": (
        "Trial Inclusion/Exclusion",
        "Eligibility content such as inclusion and exclusion categories, criterion identifiers, and criterion text.",
    ),
    "TS": (
        "Trial Summary",
        "Core study-level summary fields and registry-style metadata that feed the TS domain.",
    ),
    "TS Parameters": (
        "Trial Summary Parameters",
        "Parameter-style trial summary mappings spanning study design, roles, interventions, population, objectives, endpoints, and related metadata.",
    ),
}
SECTION_SUMMARY_EXCLUDES = {"TS Parameters"}


@dataclass(frozen=True)
class MappingRecord:
    sheet_name: str
    section_name: str
    sdtm_variable: str
    sdtm_label: str
    sdtm_type: str
    sdtm_role: str
    sdtm_core: str
    usdm_class: str
    usdm_attribute: str
    overall_notes: str
    target_path: str
    target_notes: str
    condition_path: str
    condition_notes: str
    mapping_name: str
    mapping_category: str
    status: str
    change_flag: str
    parameter_name: str = ""


@dataclass(frozen=True)
class ClassSlotMapping:
    slot_name: str
    row_number: int
    record: MappingRecord


def _clean_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).replace("\xa0", " ").strip()


def _natural_sort_key(value: str) -> list[int | str]:
    return [int(part) if part.isdigit() else part.lower() for part in re.split(r"(\d+)", value)]


def _slugify(text: str) -> str:
    text = re.sub(r"[^a-z0-9]+", "-", text.lower()).strip("-")
    return text or "section"


def _class_doc_exists(docs_path: Path, class_name: str) -> bool:
    return bool(class_name) and (docs_path / CLASS_DOCS_DIRNAME / f"{class_name}.md").exists()


def _collect_workbook_metadata(workbook_path: Path) -> tuple[list[str], list[MappingRecord]]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)

    section_order = [name for name in workbook.sheetnames if name != "Readme"]
    records: list[MappingRecord] = []

    for sheet_name in section_order:
        worksheet = workbook[sheet_name]
        rows = worksheet.iter_rows(values_only=True)
        next(rows, None)
        header = [_clean_text(value) for value in next(rows, ())]
        header_index = {name: index for index, name in enumerate(header) if name}

        if sheet_name == "TS Parameters":
            variable_column = "TSPARAMCD"
            label_column = "TSPARAM"
        else:
            variable_column = "Variable Name"
            label_column = "Variable Label"

        def get_value(row_values: list[str], column_name: str) -> str:
            index = header_index.get(column_name)
            if index is None or index >= len(row_values):
                return ""
            return row_values[index]

        for row in rows:
            row_values = [_clean_text(value) for value in row]
            if not any(row_values):
                continue

            sdtm_variable = get_value(row_values, variable_column)
            sdtm_label = get_value(row_values, label_column)
            usdm_class = get_value(row_values, "Class")

            if not sdtm_variable and not usdm_class:
                continue

            records.append(
                MappingRecord(
                    sheet_name=sheet_name,
                    section_name=sheet_name,
                    sdtm_variable=sdtm_variable,
                    sdtm_label=sdtm_label,
                    sdtm_type=get_value(row_values, "Type"),
                    sdtm_role=get_value(row_values, "Role"),
                    sdtm_core=get_value(row_values, "Core"),
                    usdm_class=usdm_class,
                    usdm_attribute=get_value(row_values, "Attribute"),
                    overall_notes=get_value(row_values, "Overall Notes"),
                    target_path=get_value(row_values, "Target Path"),
                    target_notes=get_value(row_values, "Target Notes"),
                    condition_path=get_value(row_values, "Condition Path"),
                    condition_notes=get_value(row_values, "Condition Notes"),
                    mapping_name=get_value(row_values, "Mapping Name"),
                    mapping_category=get_value(row_values, "Mapping Category"),
                    status=get_value(row_values, "Status"),
                    change_flag=get_value(row_values, "v4.0 Change"),
                    parameter_name=get_value(row_values, "TSPARAM"),
                )
            )

    return section_order, records


def _section_title(section_name: str) -> str:
    return SECTION_METADATA.get(section_name, (section_name, ""))[0]


def _section_description(section_name: str) -> str:
    return SECTION_METADATA.get(section_name, ("", "Mappings from the SDTM workbook."))[1]


def _render_section_summary(section_order: list[str], records: list[MappingRecord]) -> list[str]:
    records_by_section: dict[str, list[MappingRecord]] = defaultdict(list)
    for record in records:
        records_by_section[record.section_name].append(record)

    lines = [
        "## Explore by SDTM Section",
        "",
        '<div class="sdtm-section-grid">',
    ]

    for section_name in section_order:
        if section_name in SECTION_SUMMARY_EXCLUDES:
            continue
        title = _section_title(section_name)
        description = _section_description(section_name)
        lines.extend(
            [
                f'<a class="sdtm-section-card" href="{_slugify(section_name)}/index.html">',
                f'  <span class="sdtm-section-code">{html.escape(section_name)}</span>',
                f'  <span class="sdtm-section-title">{html.escape(title)}</span>',
                f'  <span class="sdtm-section-description">{html.escape(description)}</span>',
                '</a>',
            ]
        )

    lines.append("</div>")
    lines.append("")
    return lines


def _format_variable_display(record: MappingRecord) -> str:
    if record.section_name == "TS Parameters" and record.sdtm_variable:
        return f"TSVAL (TSPARMCD = '{record.sdtm_variable}')"
    return record.sdtm_variable or "-"


def _table_primary_column_header(section_name: str) -> str:
    if section_name == "TS Parameters":
        return "Parameter"
    return "Variable"


def _table_primary_column_value(record: MappingRecord) -> str:
    if record.section_name == "TS Parameters":
        return record.sdtm_variable or record.parameter_name or ""
    return record.sdtm_variable or ""


def _detail_page_title(section_name: str, record: MappingRecord) -> str:
    if section_name == "TS Parameters":
        parameter_code = record.sdtm_variable or record.parameter_name or "UNKNOWN"
        return f"TSPARMCD = '{parameter_code}'"
    variable_name = record.sdtm_variable or record.parameter_name or record.usdm_class or "MAPPING"
    return f"{section_name}.{variable_name}"


def _detail_directory_name(section_name: str) -> str:
    return f"{_slugify(section_name)}{DETAILS_PAGE_SUFFIX}"


def _detail_page_slug(record: MappingRecord, row_number: int) -> str:
    variable_slug = _slugify(record.sdtm_variable or record.parameter_name or record.usdm_class or "mapping")
    return f"{row_number:03d}-{variable_slug}"


def _detail_page_output_href(section_name: str, record: MappingRecord, row_number: int) -> str:
    return f"../{_detail_directory_name(section_name)}/{_detail_page_slug(record, row_number)}/index.html"


def _render_clickable_cell(content: str, href: str) -> str:
    if not content:
        return "<td></td>"
    return f'<td><a class="sdtm-row-link" href="{href}">{content}</a></td>'


def _render_clickable_row_open(detail_href: str) -> str:
    escaped_href = html.escape(detail_href, quote=True)
    return (
        f'<tr class="sdtm-clickable-row" tabindex="0" role="link" '
        f'onclick="window.location.href=\'{escaped_href}\'" '
        f'onkeydown="if (event.key === \'Enter\' || event.key === \' \') '
        f'{{ event.preventDefault(); window.location.href=\'{escaped_href}\'; }}">'
    )


def _render_usdm_class_html(record: MappingRecord, docs_path: Path, href_prefix: str = "../../../classes/") -> str:
    class_name = record.usdm_class or ""
    display_text = html.escape(class_name)
    if _class_doc_exists(docs_path, record.usdm_class):
        return f'<a href="{href_prefix}{record.usdm_class}/">{display_text}</a>'
    return display_text


def _render_usdm_class_text(record: MappingRecord) -> str:
    return html.escape(record.usdm_class or "")


def _render_usdm_class_table_cell(record: MappingRecord, docs_path: Path, detail_href: str) -> str:
    if _class_doc_exists(docs_path, record.usdm_class):
        return (
            '<td>'
            f'<a class="sdtm-class-link" href="../../classes/{record.usdm_class}/" '
            'onclick="event.stopPropagation()">'
            f"{html.escape(record.usdm_class)}"
            "</a>"
            "</td>"
        )
    if not record.usdm_class:
        return "<td></td>"
    return _render_clickable_cell(_render_usdm_class_text(record), detail_href)


def _render_target_path_text(record: MappingRecord) -> str:
    return html.escape(record.target_path or record.usdm_attribute or record.usdm_class or "")


def _render_mapping_insight(record: MappingRecord) -> str:
    detail_blocks: list[str] = []
    if record.target_notes:
        detail_blocks.append(f"<p><strong>Target Notes:</strong> {html.escape(record.target_notes)}</p>")
    if record.overall_notes:
        detail_blocks.append(f"<p>{html.escape(record.overall_notes)}</p>")
    if record.condition_notes:
        detail_blocks.append(f"<p><strong>Condition Notes:</strong> {html.escape(record.condition_notes)}</p>")
    if record.condition_path:
        detail_blocks.append(f"<p><strong>Condition Path:</strong> <code>{html.escape(record.condition_path)}</code></p>")

    if not detail_blocks:
        return "-"

    return '<div class="sdtm-rule-list">' + "".join(detail_blocks) + "</div>"


def _render_metadata_list(record: MappingRecord, docs_path: Path) -> str:
    items: list[tuple[str, str]] = []
    label_value = record.sdtm_label or record.parameter_name or ""
    class_value = _render_usdm_class_html(record, docs_path)
    path_value = _render_target_path_text(record)

    if label_value:
        items.append(("Label", html.escape(label_value)))
    if class_value:
        items.append(("USDM Class", class_value))
    if path_value:
        items.append(("USDM Path", f"<code>{path_value}</code>"))

    lines = ['<div class="sdtm-detail-card">']
    for label, value in items:
        lines.append(f'  <p class="sdtm-detail-meta"><strong>{label}:</strong> {value}</p>')
    lines.append("</div>")
    return "\n".join(lines)


def _render_domain_table_html(section_name: str, docs_path: Path, records: list[MappingRecord]) -> str:
    rows: list[str] = [
        '<table class="sdtm-mapping-table">',
        "<thead>",
        "<tr>",
        f"<th>{_table_primary_column_header(section_name)}</th>",
        "<th>Label</th>",
        "<th>USDM Class</th>",
        "<th>USDM Path</th>",
        "</tr>",
        "</thead>",
        "<tbody>",
    ]

    for row_number, record in enumerate(records, start=1):
        detail_href = _detail_page_output_href(record.section_name, record, row_number)
        primary_value = _table_primary_column_value(record)
        primary_content = f"<code>{html.escape(primary_value)}</code>" if primary_value else ""
        label_value = record.sdtm_label or record.parameter_name or ""
        rows.extend(
            [
                _render_clickable_row_open(detail_href),
                _render_clickable_cell(primary_content, detail_href),
                _render_clickable_cell(html.escape(label_value), detail_href) if label_value else "<td></td>",
                _render_usdm_class_table_cell(record, docs_path, detail_href),
                _render_clickable_cell(_render_target_path_text(record), detail_href),
                "</tr>",
            ]
        )

    rows.extend(["</tbody>", "</table>"])
    return "\n".join(rows)


def _render_detail_page(section_name: str, docs_path: Path, record: MappingRecord) -> str:
    mapping_insight = _render_mapping_insight(record)
    lines = [
        f"# {_detail_page_title(section_name, record)}",
        "",
        _render_metadata_list(record, docs_path),
        "",
        "## Mapping Details",
        "",
        mapping_insight if mapping_insight != "-" else "No additional mapping details.",
        "",
    ]

    return "\n".join(lines)


def _render_domain_page(section_name: str, docs_path: Path, records: list[MappingRecord]) -> str:
    lines = [
        f"# {_section_title(section_name)} ({section_name})",
        "",
        _section_description(section_name),
        "",
    ]

    if section_name == "TS":
        lines.extend(
            [
                "For parameter-level trial summary mappings, open [Trial Summary Parameters (TS Parameters)](ts-parameters.md).",
                "",
            ]
        )

    lines.extend(
        [
            "## Variable Map",
            "",
            "Open any row to view the full mapping details and notes on its details page.",
            "",
            _render_domain_table_html(section_name, docs_path, records),
            "",
        ]
    )

    return "\n".join(lines)


def _normalize_slot_token(value: str) -> str:
    return value.replace(" ", "").strip().lstrip("@")


def _extract_slot_candidate_from_target_path(record: MappingRecord) -> str:
    tokens = [_normalize_slot_token(token) for token in record.target_path.split("/") if token.strip()]
    class_name = _normalize_slot_token(record.usdm_class)

    for index, token in enumerate(tokens[:-1]):
        if token.lower() != class_name.lower():
            continue
        next_token = tokens[index + 1]
        if next_token.startswith("@"):
            return next_token[1:]

    return ""


def _resolve_class_slot_name(record: MappingRecord, class_slot_names: set[str]) -> str:
    normalized_slots = {_normalize_slot_token(slot_name).lower(): slot_name for slot_name in class_slot_names}
    candidates = [
        _extract_slot_candidate_from_target_path(record),
        _normalize_slot_token(record.usdm_attribute),
    ]

    for candidate in candidates:
        if not candidate:
            continue

        candidate_key = candidate.lower()
        if candidate_key in normalized_slots:
            return normalized_slots[candidate_key]

        candidate_id_key = f"{candidate_key}id"
        if candidate_id_key in normalized_slots:
            return normalized_slots[candidate_id_key]

    return ""


def _build_class_slot_mappings(schema_path: Path, records: list[MappingRecord]) -> dict[str, list[ClassSlotMapping]]:
    if SchemaView is None or not schema_path.exists():
        return {}

    schema_view = SchemaView(str(schema_path))
    class_slot_names: dict[str, set[str]] = {}
    mappings_by_class: dict[str, list[ClassSlotMapping]] = defaultdict(list)
    seen: set[tuple[str, str, str, str]] = set()

    records_by_section: dict[str, list[MappingRecord]] = defaultdict(list)
    for record in records:
        records_by_section[record.section_name].append(record)

    for section_records in records_by_section.values():
        for row_number, record in enumerate(section_records, start=1):
            if not record.usdm_class or not record.target_path:
                continue

            class_name = record.usdm_class
            if class_name not in class_slot_names:
                class_definition = schema_view.get_class(class_name)
                if class_definition is None:
                    class_slot_names[class_name] = set()
                else:
                    class_slot_names[class_name] = {
                        slot.name for slot in schema_view.class_induced_slots(class_name)
                    }

            slot_name = _resolve_class_slot_name(record, class_slot_names[class_name])
            if not slot_name:
                continue

            mapping_key = (
                class_name,
                slot_name,
                record.section_name,
                record.sdtm_variable or record.parameter_name or record.sdtm_label,
            )
            if mapping_key in seen:
                continue
            seen.add(mapping_key)

            mappings_by_class[class_name].append(
                ClassSlotMapping(slot_name=slot_name, row_number=row_number, record=record)
            )

    return mappings_by_class


def _class_mapping_domain_output_href(section_name: str) -> str:
    return f"../../{DOMAIN_PAGES_DIRNAME}/{_slugify(section_name)}/"


def _class_mapping_detail_output_href(record: MappingRecord, row_number: int) -> str:
    return (
        f"../../{DOMAIN_PAGES_DIRNAME}/"
        f"{_detail_directory_name(record.section_name)}/"
        f"{_detail_page_slug(record, row_number)}/"
    )


def _render_class_mapping_section(class_name: str, mappings: list[ClassSlotMapping]) -> str:
    lines = [
        "<!-- sdtm-class-mappings:start -->",
        "## SDTM Trial Domain Mappings",
        "",
        "This class has slot-level mappings to the following SDTM trial domain items.",
        "",
        '<table class="sdtm-class-mapping-table">',
        "<thead>",
        "<tr>",
        "<th>Domain</th>",
        "<th>SDTM Item</th>",
        "<th>Class Slot</th>",
        "<th>Label</th>",
        "</tr>",
        "</thead>",
        "<tbody>",
    ]

    for mapping in mappings:
        record = mapping.record
        item_value = record.sdtm_variable or record.parameter_name or ""
        item_label = html.escape(record.sdtm_label or record.parameter_name or "")
        lines.extend(
            [
                "<tr>",
                (
                    f'<td><a href="{_class_mapping_domain_output_href(record.section_name)}">'
                    f"{html.escape(record.section_name)}</a></td>"
                ),
                (
                    f'<td><a href="{_class_mapping_detail_output_href(record, mapping.row_number)}">'
                    f"<code>{html.escape(item_value)}</code></a></td>"
                ),
                (
                    f'<td><a href="../../slots/{mapping.slot_name}/">'
                    f"{html.escape(mapping.slot_name)}</a></td>"
                ),
                f"<td>{item_label}</td>",
                "</tr>",
            ]
        )

    lines.extend(["</tbody>", "</table>", "", "<!-- sdtm-class-mappings:end -->", ""])
    return "\n".join(lines)


def _insert_class_mapping_section(class_doc_text: str, section_text: str) -> str:
    start_marker = "<!-- sdtm-class-mappings:start -->"
    end_marker = "<!-- sdtm-class-mappings:end -->"

    if start_marker in class_doc_text and end_marker in class_doc_text:
        pattern = re.compile(
            rf"{re.escape(start_marker)}.*?{re.escape(end_marker)}\n?",
            re.DOTALL,
        )
        return pattern.sub(section_text, class_doc_text, count=1)

    insertion_markers = [
        "\n## Usages\n",
        "\n## Identifier and Mapping Information\n",
        "\n## Examples\n",
        "\n## LinkML Source\n",
    ]
    for marker in insertion_markers:
        if marker in class_doc_text:
            return class_doc_text.replace(marker, f"\n{section_text}{marker}", 1)

    return class_doc_text.rstrip() + "\n\n" + section_text


def inject_class_page_sdtm_sections(
    schema_path: Path,
    docs_path: Path,
    records: list[MappingRecord],
) -> None:
    if not records:
        return

    class_docs_path = docs_path / CLASS_DOCS_DIRNAME
    if not class_docs_path.exists():
        return

    mappings_by_class = _build_class_slot_mappings(schema_path, records)
    for class_name, mappings in mappings_by_class.items():
        class_doc_path = class_docs_path / f"{class_name}.md"
        if not class_doc_path.exists():
            continue

        original_text = class_doc_path.read_text(encoding="utf-8")
        updated_text = _insert_class_mapping_section(
            original_text,
            _render_class_mapping_section(class_name, mappings),
        )
        if updated_text != original_text:
            class_doc_path.write_text(updated_text, encoding="utf-8")


def _write_domain_pages(docs_path: Path, section_order: list[str], records: list[MappingRecord]) -> None:
    records_by_section: dict[str, list[MappingRecord]] = defaultdict(list)
    for record in records:
        records_by_section[record.section_name].append(record)

    domain_docs_path = docs_path / DOMAIN_PAGES_DIRNAME
    domain_docs_path.mkdir(parents=True, exist_ok=True)

    for section_name in section_order:
        section_records = records_by_section.get(section_name, [])
        page_path = domain_docs_path / f"{_slugify(section_name)}.md"
        page_path.write_text(
            _render_domain_page(section_name, docs_path, section_records),
            encoding="utf-8",
        )

        detail_docs_path = domain_docs_path / _detail_directory_name(section_name)
        detail_docs_path.mkdir(parents=True, exist_ok=True)
        for row_number, record in enumerate(section_records, start=1):
            detail_page_path = detail_docs_path / f"{_detail_page_slug(record, row_number)}.md"
            detail_page_path.write_text(
                _render_detail_page(section_name, docs_path, record),
                encoding="utf-8",
            )


def write_sdtm_mapping_docs(workbook_path: Path, docs_path: Path) -> list[MappingRecord]:
    if load_workbook is None:
        return []

    if not workbook_path.exists():
        return []

    section_order, records = _collect_workbook_metadata(workbook_path)
    _write_domain_pages(docs_path, section_order, records)
    return records
